import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from jinja2 import Environment, FileSystemLoader, select_autoescape

from app import db
from config.settings import BASE_DIR
from core.ai_analyzer import analyze_attack_chain, analyze_vulnerability
from database.db_operation import get_vulnerabilities_by_task_id
from database.models import ScanTask, Vulnerability

logger = logging.getLogger("report")


def _get_task_and_vulns(task_id: str) -> Tuple[ScanTask | None, List[Vulnerability]]:
    """
    根据 task_id 获取任务信息和漏洞列表。

    :param task_id: 扫描任务 ID（ScanTask.task_id）
    :return: (任务实例或 None, 漏洞列表)
    """
    task = ScanTask.query.filter_by(task_id=task_id).first()
    if not task:
        return None, []

    vulns = (
        Vulnerability.query.filter_by(task_id=task_id)
        .order_by(Vulnerability.risk_level.desc(), Vulnerability.create_time.asc())
        .all()
    )
    return task, vulns


def _ensure_dir(path: str) -> None:
    """
    确保目录存在。

    :param path: 目录路径
    """
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


def _run_ai_analysis_for_task(task_id: str) -> None:
    """
    在报告生成阶段触发 AI 分析。

    - 对任务下的每个漏洞执行一次 analyze_vulnerability（带缓存判断）；
    - 再执行一次攻击链分析，将整体结论写入各条漏洞记录；
    - 任何 AI 相关异常都只记录日志，不阻断报告生成。
    """
    try:
        vulns = get_vulnerabilities_by_task_id(task_id)
        if not vulns:
            logger.info("AI 分析跳过：任务下无漏洞记录 task_id=%s", task_id)
            return

        for v in vulns:
            try:
                analyze_vulnerability(vuln_id=v.id, vuln_data=v)
            except Exception as exc:  # noqa: BLE001
                logger.warning("单条漏洞 AI 分析失败，已跳过: vuln_id=%s error=%s", v.id, exc)

        try:
            analyze_attack_chain(task_id)
        except Exception as exc:  # noqa: BLE001
            logger.warning("攻击链 AI 分析失败，已使用降级说明: task_id=%s error=%s", task_id, exc)
    except Exception as exc:  # noqa: BLE001
        logger.warning("AI 分析阶段整体异常，报告将使用降级信息: task_id=%s error=%s", task_id, exc)


def _risk_fill(level: str) -> PatternFill:
    """
    根据风险等级返回 Excel 单元格填充颜色。
    """
    level_l = (level or "").lower()
    if level_l == "high":
        color = "FFC7CE"  # 红色
    elif level_l == "medium":
        color = "FFEB9C"  # 黄色
    elif level_l == "low":
        color = "C6EFCE"  # 绿色
    else:
        color = "D9D9D9"  # 灰色
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def generate_excel_report(task_id: str, output_path: str) -> Tuple[bool, str]:
    """
    生成 Excel 漏洞报告。

    报告内容包括：漏洞基础信息与 AI 分析结果，表头加粗、风险等级着色、自动列宽。

    :param task_id: 扫描任务 ID
    :param output_path: 输出文件完整路径（.xlsx）
    :return: (是否成功, 提示信息)
    """
    try:
        # 在生成报告前先运行一次 AI 分析，用于丰富报告内容
        _run_ai_analysis_for_task(task_id)

        task, vulns = _get_task_and_vulns(task_id)
        if not task:
            return False, "任务不存在，无法生成报告。"

        wb = Workbook()
        ws = wb.active
        ws.title = "漏洞详情"

        headers = [
            "漏洞 ID",
            "任务 ID",
            "目标域名/IP+端口",
            "漏洞名称",
            "风险等级",
            "URL",
            "参数",
            "扫描工具",
            "工具描述",
            "AI 攻击原理分析",
            "AI PoC 建议（思路）",
            "AI 定制化修复建议",
            "AI 攻击链风险评估",
            "记录时间",
        ]

        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_align

        if not vulns:
            # 仍生成“空报告”，保证流程跑通并可下载
            ws.append(
                [
                    "",
                    task_id,
                    task.domain,
                    "未发现漏洞（或扫描模块部分未启用）",
                    "info",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
                ]
            )
        else:
            for v in vulns:
                row = [
                    v.id,
                    v.task_id,
                    task.domain,
                    v.vuln_name,
                    v.risk_level,
                    v.url,
                    v.param or "",
                    v.tool,
                    v.description or "",
                    v.attack_principle or "",
                    v.poc_suggestion or "",
                    v.custom_fix or "",
                    v.attack_chain_risk or "",
                    v.create_time.strftime("%Y-%m-%d %H:%M:%S"),
                ]
                ws.append(row)

        # 风险等级着色 + 自动列宽
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            risk_cell = row[4]  # 第 5 列为风险等级
            risk_cell.fill = _risk_fill(str(risk_cell.value))
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx in range(1, len(headers) + 1):
            column = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[column]:
                try:
                    value_len = len(str(cell.value)) if cell.value is not None else 0
                    if value_len > max_length:
                        max_length = value_len
                except Exception:
                    continue
            ws.column_dimensions[column].width = min(max_length + 2, 60)

        _ensure_dir(os.path.dirname(output_path))
        wb.save(output_path)

        file_size = os.path.getsize(output_path)
        logger.info(
            "Excel 报告生成成功: task_id=%s path=%s size=%s bytes",
            task_id,
            output_path,
            file_size,
        )
        return True, f"Excel 报告生成成功（大小 {file_size} 字节）。"
    except PermissionError as exc:
        logger.exception("Excel 报告写入权限不足: %s", exc)
        return False, "Excel 报告写入失败（权限不足），请检查目标目录权限。"
    except Exception as exc:  # noqa: BLE001
        logger.exception("Excel 报告生成失败: %s", exc)
        return False, "Excel 报告生成失败，请稍后重试。"


def _render_pdf_html(task: ScanTask, vulns: List[Vulnerability]) -> str:
    """
    使用 Jinja2 渲染 PDF 报告 HTML。

    :param task: 任务实例
    :param vulns: 漏洞列表
    :return: 渲染后的 HTML 字符串
    """
    templates_dir = os.path.join(BASE_DIR, "app", "templates")
    env = Environment(
        loader=FileSystemLoader(templates_dir, encoding="utf-8"),
        autoescape=select_autoescape(["html", "xml"]),
    )
    template = env.get_template("report/pdf_report_template.html")

    # 漏洞统计
    total = len(vulns)
    stats: Dict[str, int] = {"high": 0, "medium": 0, "low": 0, "info": 0}
    for v in vulns:
        level = (v.risk_level or "info").lower()
        level = "info" if level not in stats else level
        stats[level] += 1

    context: Dict[str, Any] = {
        "task": task,
        "vulns": vulns,
        "stats": stats,
        "total": total,
        "generated_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
    }
    return template.render(**context)


def generate_pdf_report(task_id: str, output_path: str) -> Tuple[bool, str]:
    """
    生成 PDF 渗透测试报告。

    使用 Jinja2 模板渲染 HTML，再通过 WeasyPrint 转为 PDF。

    :param task_id: 扫描任务 ID
    :param output_path: 输出文件完整路径（.pdf）
    :return: (是否成功, 提示信息)
    """
    try:
        # 在生成报告前先运行一次 AI 分析，用于丰富报告内容
        _run_ai_analysis_for_task(task_id)

        task, vulns = _get_task_and_vulns(task_id)
        if not task:
            return False, "任务不存在，无法生成报告。"

        html_content = _render_pdf_html(task, vulns)

        try:
            from weasyprint import HTML as WeasyHTML
        except OSError as exc:
            logger.warning("WeasyPrint 系统依赖未安装（如 GTK/Cairo），PDF 无法生成: %s", exc)
            return False, "PDF 报告需要 WeasyPrint 系统依赖（GTK/Cairo），当前环境未满足，请安装后重试或仅下载 Excel 报告。"

        _ensure_dir(os.path.dirname(output_path))
        base_url = os.path.join(BASE_DIR, "app", "templates")
        WeasyHTML(string=html_content, base_url=base_url).write_pdf(target=output_path)

        file_size = os.path.getsize(output_path)
        logger.info(
            "PDF 报告生成成功: task_id=%s path=%s size=%s bytes",
            task_id,
            output_path,
            file_size,
        )
        return True, f"PDF 报告生成成功（大小 {file_size} 字节）。"
    except FileNotFoundError as exc:
        logger.exception("PDF 模板文件缺失: %s", exc)
        return False, "PDF 模板文件缺失，请检查模板路径。"
    except PermissionError as exc:
        logger.exception("PDF 报告写入权限不足: %s", exc)
        return False, "PDF 报告写入失败（权限不足），请检查目标目录权限。"
    except Exception as exc:  # noqa: BLE001
        logger.exception("PDF 报告生成失败: %s", exc)
        return False, "PDF 报告生成失败，请稍后重试。"


__all__ = [
    "generate_excel_report",
    "generate_pdf_report",
]

